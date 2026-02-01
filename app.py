import streamlit as st
import google.generativeai as genai
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import io
from PIL import Image
import concurrent.futures

# ==========================================
# 【ユーザー設定エリア: 過去の文体学習】
# ここにあなたの過去の感想文をコピペしてください。
# AIはこの文章の「書き出し」「熱量」「言葉選び」を真似します。
# ==========================================
PAST_REVIEWS = """
（例：過去の感想文）
今月の致知を読んで、特に「逆境こそが人を育てる」という言葉が胸に刺さりました。
日々の税理士補助業務において、繁忙期にはつい愚痴が出そうになりますが、
それは自分の魂を磨く砥石なのだと気づかされました。
お客様の試算表を作る作業一つとっても、そこに魂を込めること。
それがプロフェッショナルとしての流儀だと感じます。
"""

# ==========================================
# ページ設定
# ==========================================
st.set_page_config(page_title="致知読書感想文アプリ v4.0", layout="wide", page_icon="📖")
st.title("📖 致知読書感想文作成アプリ (並列OCR & 記事選択版)")
st.caption("Step 1: 並列OCR(Gemini 3/2.5) → Step 2: 記事選択・執筆・壁打ち → Step 3: Excel出力")

# Excel書き込み設定（厳守）
EXCEL_START_ROW = 9
CHARS_PER_LINE = 40

# ==========================================
# API設定 & セッション初期化
# ==========================================
try:
    openai_key = st.secrets.get("OPENAI_API_KEY")
    client = OpenAI(api_key=openai_key) if openai_key else None
    
    google_key = st.secrets.get("GOOGLE_API_KEY")
    if google_key:
        genai.configure(api_key=google_key)
        
except Exception as e:
    st.error(f"API設定エラー: {e}")

# セッション状態の初期化
if "ocr_results" not in st.session_state:
    # 記事ごとのOCR結果を辞書で保存（混ざらないようにする）
    st.session_state.ocr_results = {"main": "", "sub1": "", "sub2": ""}
if "current_draft" not in st.session_state:
    st.session_state.current_draft = ""
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "selected_article_key" not in st.session_state:
    st.session_state.selected_article_key = "main"

# ==========================================
# 関数定義
# ==========================================
def split_text(text, chunk_size):
    """Excel用にテキストを指定文字数で分割"""
    if not text:
        return []
    clean_text = text.replace('\n', '　')
    return [clean_text[i:i+chunk_size] for i in range(0, len(clean_text), chunk_size)]

def process_ocr_task(label, files, model_id):
    """
    【並列処理用】OCR関数
    指定された画像ファイル群をGeminiで解析してテキストを返す
    """
    if not files:
        return ""
    
    try:
        gemini_inputs = []
        # プロンプト：読み取り順序の指定（厳格なルール）
        system_prompt = (
            "あなたはOCRエンジンです。雑誌『致知』の「上下2段組み」画像を読み取ります。\n"
            "【厳守ルール】\n"
            "1. 画像を「上段」と「下段」に分けて認識する。\n"
            "2. まず【上段】の文章を右から左へ読む。\n"
            "3. 次に【下段】の文章を右から左へ読む。\n"
            "4. ※絶対に左側の段を上から下へ一気に読まないこと。\n"
            "5. 出力形式: [画像番号] <上段>... <下段>..."
        )
        gemini_inputs.append(system_prompt)
        
        for img_file in files:
            img_file.seek(0)
            image = Image.open(img_file).convert("RGB")
            gemini_inputs.append(image)
        
        # モデル生成と実行
        model = genai.GenerativeModel(model_id)
        response = model.generate_content(gemini_inputs)
        return response.text
        
    except Exception as e:
        return f"[エラー: {label}の解析失敗: {e}]"

def generate_draft(article_text, chat_context, target_len):
    """感想文執筆関数 (文体模倣 + 壁打ち反映)"""
    system_prompt = (
        "あなたは税理士事務所の職員です。\n"
        "これから雑誌『致知』の読書感想文（社内木鶏会用）を作成します。\n"
        "以下の【ユーザーの過去の感想文】を分析し、"
        "**「文体」「書き出しの癖」「精神的な熱量」「業務（巡回監査・決算など）への結びつけ方」**を模倣してください。"
    )
    user_content = (
        f"【今回選択した記事のOCRデータ】\n{article_text}\n\n"
        f"【ユーザーの過去の感想文（スタイル見本）】\n{PAST_REVIEWS}\n\n"
        f"【壁打ちチャットでの打ち合わせ内容（ここでのエピソードを必ず盛り込むこと）】\n{chat_context}\n\n"
        "【執筆条件】\n"
        f"- 文字数：{target_len}文字前後\n"
        "- 文体：「です・ます」調\n"
        "- 段落ごとに改行を入れること。\n"
        "- 構成：①記事の引用 ②自分の業務エピソード ③今後の決意"
    )
    
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_content}],
        temperature=0.7
    )
    return response.choices[0].message.content

# ==========================================
# サイドバー設定
# ==========================================
with st.sidebar:
    st.header("⚙️ 設定")
    uploaded_template = st.file_uploader("感想文フォーマット(.xlsx)", type=["xlsx"])
    target_length = st.selectbox("目標文字数", [300, 400, 500, 600, 700, 800], index=1)
    
    st.markdown("---")
    st.caption("🔧 OCRモデル設定")
    
    # 【重要】ご指定のモデルIDをデフォルト値に設定
    st.write("メイン記事用 (精度重視):")
    model_main = st.text_input("メインModel ID", value="gemini-3-flash-preview")
    
    st.write("サブ記事用 (速度重視):")
    model_sub = st.text_input("サブModel ID", value="gemini-2.5-flash-lite")
    st.caption("※API側でモデルが有効か確認してください")

    if st.button("🗑️ リセット"):
        for key in st.session_state.keys():
            del st.session_state[key]
        st.rerun()

# ==========================================
# メイン画面 (タブ構成)
# ==========================================
tab1, tab2, tab3 = st.tabs(["1️⃣ 画像解析 (並列処理)", "2️⃣ 記事選択 & 執筆", "3️⃣ Excel出力"])

# ------------------------------------------------------------------
# Tab 1: 並列OCR処理
# ------------------------------------------------------------------
with tab1:
    st.subheader("Step 1. 記事画像の読み込み (並列処理)")
    st.info("メイン記事とサブ記事を並列で高速解析します。")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("#### 📂 メイン記事")
        files_main = st.file_uploader("画像を選択", type=['png', 'jpg', 'jpeg'], accept_multiple_files=True, key="f1")
    with col2:
        st.markdown("#### 📂 記事2")
        files_sub1 = st.file_uploader("画像を選択", type=['png', 'jpg', 'jpeg'], accept_multiple_files=True, key="f2")
    with col3:
        st.markdown("#### 📂 記事3")
        files_sub2 = st.file_uploader("画像を選択", type=['png', 'jpg', 'jpeg'], accept_multiple_files=True, key="f3")

    # 解析ボタン
    if st.button("🚀 全記事を一括解析 (並列スタート)", type="primary"):
        if not (files_main or files_sub1 or files_sub2):
            st.error("画像が選択されていません。")
        else:
            with st.spinner("指定されたモデルで3つの記事を同時に解析中..."):
                # 並列処理の実行 (Concurrent Futures)
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    # タスクの登録 
                    # メインは model_main (gemini-3-flash-preview)
                    # サブは model_sub (gemini-2.5-flash-lite 等)
                    future_main = executor.submit(process_ocr_task, "メイン記事", files_main, model_main)
                    future_sub1 = executor.submit(process_ocr_task, "記事2", files_sub1, model_sub)
                    future_sub2 = executor.submit(process_ocr_task, "記事3", files_sub2, model_sub)
                    
                    # 結果の取得（各スレッドの完了を待機）
                    st.session_state.ocr_results["main"] = future_main.result()
                    st.session_state.ocr_results["sub1"] = future_sub1.result()
                    st.session_state.ocr_results["sub2"] = future_sub2.result()
                
                st.success("✅ 全ての解析が完了しました！ '2️⃣ 記事選択 & 執筆' タブへ進んでください。")

    # 結果確認用 (デバッグ)
    with st.expander("OCR解析結果を確認する"):
        st.markdown("###### メイン記事")
        st.text_area("Main", st.session_state.ocr_results["main"], height=100)
        st.markdown("###### 記事2")
        st.text_area("Sub1", st.session_state.ocr_results["sub1"], height=100)
        st.markdown("###### 記事3")
        st.text_area("Sub2", st.session_state.ocr_results["sub2"], height=100)

# ------------------------------------------------------------------
# Tab 2: 記事選択 & 執筆 & 壁打ち (Core Feature)
# ------------------------------------------------------------------
with tab2:
    st.subheader("Step 2. 執筆対象の選択と壁打ち")
    
    # ==========================
    # 1. 記事の選択 (Radio Button)
    # ==========================
    st.markdown("##### どの記事の感想文を書きますか？")
    
    # データが存在する選択肢のみを有効化
    options_map = {"main": "メイン記事", "sub1": "記事2", "sub2": "記事3"}
    valid_options = [k for k, v in st.session_state.ocr_results.items() if len(v) > 10] # 10文字以上なら有効とみなす
    
    if not valid_options:
        st.warning("OCRデータがありません。Tab 1で解析を行ってください。")
        selected_article_text = ""
    else:
        # ラジオボタン表示
        selected_key = st.radio(
            "対象記事を選択", 
            valid_options, 
            format_func=lambda x: options_map[x],
            horizontal=True
        )
        selected_article_text = st.session_state.ocr_results[selected_key]
        
        # 選択切り替え時にチャット履歴をリセットするか確認してもいいが、今回は簡易的に表示のみ切り替え
        # 記事テキストを表示
        with st.expander("選択中の記事内容を表示"):
            st.text(selected_article_text)

    st.markdown("---")

    # ==========================
    # 2. 執筆 & 壁打ち UI
    # ==========================
    col_draft, col_chat = st.columns([1, 1])

    # --- 左側：感想文ドラフト ---
    with col_draft:
        st.markdown("### 📝 感想文ドラフト")
        
        if st.button("🚀 初稿を作成する", disabled=(not selected_article_text)):
            with st.spinner("過去の文体を分析して執筆中..."):
                # チャット履歴なしで初稿作成
                draft = generate_draft(selected_article_text, "", target_length)
                st.session_state.current_draft = draft
                # 壁打ち開始のトリガーメッセージ
                st.session_state.chat_history = [{
                    "role": "assistant", 
                    "content": "初稿を作成しました！\nより良い感想文にするために、この記事に関連するあなたの具体的な体験談（業務での出来事など）を教えてください。"
                }]
                st.rerun()
        
        if st.session_state.current_draft:
            st.text_area("現在の原稿", st.session_state.current_draft, height=600, key="draft_area")
            
            st.info("👈 右側のチャットでエピソードを追加し、下のボタンで書き直せます。")
            if st.button("🔄 チャットの内容を反映して書き直す", type="primary"):
                with st.spinner("会話内容を反映してリライト中..."):
                    # チャット履歴をテキスト化して渡す
                    chat_context = "\n".join([f"{m['role']}: {m['content']}" for m in st.session_state.chat_history])
                    new_draft = generate_draft(selected_article_text, chat_context, target_length)
                    st.session_state.current_draft = new_draft
                    st.success("書き直しました！")
                    st.rerun()

    # --- 右側：壁打ちチャット ---
    with col_chat:
        st.markdown("### 💬 壁打ち (思考の深掘り)")
        
        chat_container = st.container(height=500)
        
        # 履歴表示
        for message in st.session_state.chat_history:
            with chat_container.chat_message(message["role"]):
                st.markdown(message["content"])

        # 入力フォーム
        if prompt := st.chat_input("エピソードや考えを入力..."):
            if not selected_article_text:
                st.error("先に記事を選択して初稿を作成してください。")
            else:
                st.session_state.chat_history.append({"role": "user", "content": prompt})
                with chat_container.chat_message("user"):
                    st.markdown(prompt)

                with chat_container.chat_message("assistant"):
                    with st.spinner("考え中..."):
                        # 壁打ち用プロンプト（記事内容だけを前提にする）
                        chat_system = (
                            "あなたは優秀な編集者です。\n"
                            "以下の記事内容を読んだユーザーに対して、より深い感想を引き出すための質問をしてください。\n"
                            f"【記事内容】: {selected_article_text[:500]}...\n"
                            "具体的な業務経験（成功・失敗）と感情を引き出すことに集中してください。"
                        )
                        chat_messages = [{"role": "system", "content": chat_system}] + \
                                        [{"role": m["role"], "content": m["content"]} for m in st.session_state.chat_history]
                        
                        res = client.chat.completions.create(
                            model="gpt-4o",
                            messages=chat_messages,
                            temperature=0.7
                        )
                        ai_response = res.choices[0].message.content
                        
                st.markdown(ai_response)
                st.session_state.chat_history.append({"role": "assistant", "content": ai_response})

# ------------------------------------------------------------------
# Tab 3: Excel出力
# ------------------------------------------------------------------
with tab3:
    st.subheader("Step 3. Excelへの書き出し")
    
    if not st.session_state.current_draft:
        st.warning("まだ感想文が作成されていません。")
    else:
        st.write("完成した以下のテキストをExcelに出力します。")
        st.text(st.session_state.current_draft)
        
        if uploaded_template:
            if st.button("📥 Excelを作成してダウンロード"):
                try:
                    wb = load_workbook(uploaded_template)
                    ws = wb.active
                    
                    # 以前の内容をクリア（A9セル以降）
                    for row in range(EXCEL_START_ROW, 100):
                        ws[f"A{row}"].value = None
                    
                    # 40文字区切りでリスト化
                    lines = split_text(st.session_state.current_draft, CHARS_PER_LINE)
                    
                    # A9セルから順に書き込み
                    for i, line in enumerate(lines):
                        current_row = EXCEL_START_ROW + i
                        cell = ws[f"A{current_row}"]
                        cell.value = line
                        # 書式設定（折り返さない、縮小しない、左寄せ）
                        cell.alignment = Alignment(wrap_text=False, shrink_to_fit=False, horizontal='left')
                    
                    # バッファに保存
                    out = io.BytesIO()
                    wb.save(out)
                    out.seek(0)
                    
                    st.download_button(
                        label="Excelファイルを保存",
                        data=out,
                        file_name="社内木鶏会感想文.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success("出力完了！")
                    
                except Exception as e:
                    st.error(f"Excel出力エラー: {e}")
        else:
            st.warning("テンプレートExcel（感想文フォーマット.xlsx）をサイドバーからアップロードしてください。")
